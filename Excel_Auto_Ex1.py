import time
from selenium import webdriver
import openpyxl
import subprocess
import pyautogui
#-------------------------------------------------------------------*
#システム仕様書EXCELファイルを読込み、Pythonのソースコードを出力する
#-------------------------------------------------------------------*

SPECIFIC_SRC_FILE=r'D:\auto_py_spec\spec_sheet01.xlsx'#仕様書ファイル
OUTPUT_FILE_DIR=r'D:\py_auto_out'#ソースコードファイルのディレクトリ
OUTPUT_FILE_SRC=r'D:\py_auto_out\py_src_excel.txt'#ソースコード出力ファイル

wb = openpyxl.load_workbook(SPECIFIC_SRC_FILE)
#print(wb.get_sheet_names())
sheet=wb.get_sheet_by_name("PTN-1")

pyautogui.alert('システムをスタートします')

print("#-----------------------------------------------------------------------------#")
print(" 仕様書EXCELファイル(spec_sheet01.xlsx)を読込みPythonのソースコードを出力します")
print(" 出力ファイルはpy_src_excel.txtdです")
print("#-----------------------------------------------------------------------------#")

def out_src_file(para_m):
    #global init_in_out_file
    path_w = OUTPUT_FILE_SRC
    if para_m==0:
        with open(path_w, mode='w') as f:
            imp1="import openpyxl\n"
            imp2="import pyautogui\n"
            imp3="import time\n"
            imp4="import subprocess\n"
            imp5="import os\n"
            imp6="import sys\n\n"

            imp_all=imp1+imp2+imp3+imp4+imp5+imp6

            cmnt1="#-----------------------------------------------#\n"
            cmnt2="#機能：Excel To Excel のコピー"
            cmnt3="#\n"
            cmnt4="#-----------------------------------------------#\n\n"
            cmnt_all=cmnt1+cmnt2+cmnt3+cmnt4
            start_time_delay="time.sleep(2)\n"#システムスタートまでの待ち時間
            #init_in_out_file="AAAAA"
            f.write(imp_all+cmnt_all+start_time_delay+init_in_out_file+"\n")
    elif para_m==1:
        with open(path_w, mode='a') as f:
            f.write(data_cell_cpoy)
#---------------------------------------------------------------------------#
def init():
    #print("INIT-START")
    global init_in_out_file
    global out_name
    #インプット・アウトプットの情報ゲット-------------------------------*
    in_dir = sheet["B3"].value#インプットファイルのディレクトリ名
    in_file = sheet["C3"].value#インプットファイル名
    in_name=in_dir+"\\"+in_file#フルパス名
    in_tab=sheet['G3'].value #タブ名
    out_dir = sheet["D3"].value#アウトプットファイルのディレクトリ名
    out_file = sheet["E3"].value#アウトプットファイル名
    out_name=out_dir+"\\"+out_file#フルパス名
    out_tab=sheet['H3'].value #タブ名
    #------------------------------------------------------------------*

    #init_in_out_file = init_in_out_file + "\nwb_in = openpyxl.load_workbook(" + "'" + in_name + ",data_only=True"+"'"")"
    init_in_out_file = init_in_out_file + "\nwb_in = openpyxl.load_workbook(" + "'" + in_name +"'"+",data_only=True)"
    init_in_out_file = init_in_out_file+"\n" + "sheet_in=wb_in.get_sheet_by_name(" + "'"+in_tab+"'"+")"
    init_in_out_file = init_in_out_file+"\n" + "wb_out = openpyxl.load_workbook("+ "'"+out_name+"'"+")"
    init_in_out_file = init_in_out_file+"\n"+"sheet_out=wb_out.get_sheet_by_name(" + "'"+out_tab+"'"+")"
    #print("init_in_out_file ",init_in_out_file)

    out_src_file(0)#処理データのアウトプット

#<Input→OutPut>-------------------------------------------------------------
def data_copy():
    global data_cell_cpoy
    max_r = wb['PTN-1'].max_row # 最終行の取得
    #print("max_r",max_r)
    for cntr in range(3,max_r):#3行目からスタート
        in_dat = sheet['I' + str(cntr)].value#コピー元のセルが記載されている
        out_dat = sheet['J' + str(cntr)].value#コピー先のセルが記載されている
        #print("in_dat:",in_dat)
        #print("out_dat:", out_dat)
        data_cell_cpoy = data_cell_cpoy + "\nsheet_out['" + out_dat + "'""].value" + "=sheet_in['" + in_dat + "'""].value"

    data_cell_cpoy=data_cell_cpoy+"\n\nwb_out.save('" + out_name +"')"
    wb.close()#ファイルのクローズ

    #print("data_cell_cpoy:",data_cell_cpoy)
    out_src_file(1)  # コピー内容のアウトプット

def end_proc():
    #pyautogui.alert('終了しました')
    subprocess.run('explorer {}'.format(OUTPUT_FILE_DIR))  # 終了と同時にソースコードのアウトプットフォルダを開く

init_in_out_file=""
init()
data_cell_cpoy=""
data_copy()
end_proc()

