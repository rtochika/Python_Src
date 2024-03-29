import csv
import openpyxl
import sys
import pyautogui
#-------------------------------------------------#
#ケイヒン航空の５事業所の会計システムから出力した
# PLデータ（CSV）を転記する
#-------------------------------------------------#
deb_flg =1 # 1:ON 0:OFF

#Global変数
OUTPUT_FILE = ""# 営業実績
AIR_KANT_EIGYO_=""# 関東営業
AIR_NARITA_F=""# 成田
AIR_HANED_F =""# 羽田
AIR_KANS_EIGYO_F = ""#関西営業
AIR_KANKU_F = "" # 関空
AIR_GYOSEKI_F = "" # 業績ファイル
AIR_HAIFU_F = "" # 配賦ファイル
AIR_TSUKI = "" # 計上月（営業実績のタブ指定）

#------------------------------------------------------#
#初期処理
#------------------------------------------------------#
def initial():
    global OUTPUT_FILE
    global AIR_KANT_EIGYO_F
    global AIR_NARITA_F
    global AIR_HANED_F
    global AIR_KANS_EIGYO_F
    global AIR_KANKU_F
    global AIR_GYOSEKI_F
    global AIR_HAIFU_F
    global AIR_TSUKI

    INPUT_CSV=r'C:\k-net\航空PL\kakei_air_pl_initial.txt'
    cntr=0
    with open(INPUT_CSV) as f:
        reader = csv.reader(f)
        for row in reader:
            if row[0] !="#":
                cntr=cntr+1
                if cntr==1:
                    OUTPUT_FILE = "c:/k-net/航空pl/"+row[0]  # 例：営業実績
                elif cntr == 2:
                    AIR_KANT_EIGYO_F = "c:/k-net/航空pl/"+row[0]  # 関東営業
                elif cntr == 3:
                    AIR_NARITA_F = "c:/k-net/航空pl/"+row[0]  # 成田
                elif cntr == 4:
                    AIR_HANED_F = "c:/k-net/航空pl/"+row[0]  # 羽田
                elif cntr == 5:
                    AIR_KANS_EIGYO_F = "c:/k-net/航空pl/"+row[0]  # 関西営業
                elif cntr == 6:
                    AIR_KANKU_F = "c:/k-net/航空pl/"+row[0]  # 関空
                elif cntr == 7:
                    AIR_GYOSEKI_F = "c:/k-net/航空pl/"+row[0]  # 業績
                elif cntr == 8:
                    AIR_HAIFU_F = "c:/k-net/航空pl/"+row[0]  # 配賦
                elif cntr == 9:
                    AIR_TSUKI = row[0]  # 計上月


#------------------------------------------------------#
#関東・関西営業
#------------------------------------------------------#
def exec_sales(cnt):#関東・関西営業

    if cnt==0:#関東営業
        FILE_in_sales = AIR_KANT_EIGYO_F
    else:#関西営業
        FILE_in_sales = AIR_KANS_EIGYO_F
    #sys.exit()

    AC_AIR_32111 = "32111" #航空輸出収益  混載
    AC_AIR_32112 = "32112" #航空輸出収益  一般
    AC_AIR_32113 = "32113" #航空輸出収益  入出庫料
    AC_AIR_32114 = "32114" #航空輸出収益  運送料
    AC_AIR_32116 = "32116" #航空輸出収益  通関料
    AC_AIR_32117 = "32117" #航空輸出収益  クーリエ収益(OBC)
    AC_AIR_32119 = "32119" #航空輸出収益  その他

    AC_AIR_32121 = "32121" #航空輸出費用  混載支払
    AC_AIR_32123 = "32123" #航空輸出費用  入出庫費
    AC_AIR_32124 = "32124" #航空輸出費用  委託運送費
    AC_AIR_32125 = "32125" #航空輸出費用  情報処理費  要確認"
    AC_AIR_32127 = "32127" #航空輸出費用  クーリエ費用(OBC)
    AC_AIR_32126 = "32126" #航空輸出費用  通関費
    AC_AIR_32128 = "32128" #航空輸出費用  その他"

    #値の初期化
    air_32111 = 0  # 航空輸出収益  混載
    air_32112 = 0  # 航空輸出収益  一般
    air_32113 = 0  # 航空輸出収益  入出庫料
    air_32114 = 0  # 航空輸出収益  運送料
    air_32116 = 0  # 航空輸出収益  通関料
    air_32117 = 0  # 航空輸出収益  クーリエ収益(OBC)
    air_32119 = 0  # 航空輸出収益  その他

    air_32121 = 0  # 航空輸出費用  混載支払
    air_32123 = 0  # 航空輸出費用  入出庫費
    air_32124 = 0  # 航空輸出費用  委託運送費
    air_32125 = 0  # 航空輸出費用  情報処理費  要確認"
    air_32126 = 0  # 航空輸出費用  通関費"
    air_32127 = 0  # 航空輸出費用  クーリエ費用(OBC)
    air_32128 = 0  # 航空輸出費用  その他"

    ACCOUNT_POS=0 #勘定科目の列番目
    TOGETSU_POS=4#当月金額の列番目

    #転送先セル-------------------------------------------------*
    #収入
    if cnt==0:#関東営業
        KONSAI = "F7"  #混載
        TESURY = "F8"  #手数料
        NYUSYU = "F9"  #入出庫
        UNSORO = "F10" #運送料
        SEAAIR = "F11" #SEA & AIR
        TSUKAN = "F12" #通関料
        OBCSYU = "F13" #OBC収入
        SONOTA = "F14" #その他
        #費用
        KONGEN = "F16" #混載原価
        NYUSYH = "F17" #入出庫費
        UNSOHI = "F18" #運送費
        SEAARH = "F19" #SEA & AIR
        TSUKAH = "F20" #通関費
        OBCHIY = "F21" #OBC費用
        SONOTH = "F22" #その他費用
    else:#関西営業
        KONSAI = "Z7"  #混載
        TESURY = "Z8"  #手数料
        NYUSYU = "Z9"  #入出庫
        UNSORO = "Z10" #運送料
        SEAAIR = "Z11" #SEA & AIR
        TSUKAN = "Z12" #通関料
        OBCSYU = "Z13" #OBC収入
        SONOTA = "Z14" #その他
        #費用
        KONGEN = "Z16" #混載原価
        NYUSYH = "Z17" #入出庫費
        UNSOHI = "Z18" #運送費
        SEAARH = "Z19" #SEA & AIR
        TSUKAH = "Z20" #通関費
        OBCHIY = "Z21" #OBC費用
        SONOTH = "Z22" #その他費用

    #----------------------------------------------------*
    cntr=0
    with open(FILE_in_sales) as f:
        #next(f)#１行目がブランクだとエラーになるので苦肉の策（２行目から読む）
        reader = csv.reader(f)
        for row in reader:
            cntr=cntr+1
            if(AC_AIR_32111 in row[ACCOUNT_POS]):#航空輸出収益  混載 合致する
                air_32111=air_32111+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32112 in row[ACCOUNT_POS]):#航空輸出収益  一般 合致する:
                air_32112=air_32112+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32113 in row[ACCOUNT_POS]):#航空輸出収益  入出庫料
                air_32113=air_32113+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32114 in row[ACCOUNT_POS]):#航空輸出収益  運送料
                air_32114=air_32114+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32116 in row[ACCOUNT_POS]):#航空輸出収益  通関料
                air_32116=air_32116+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32117 in row[ACCOUNT_POS]):#航空輸出収益  クーリエ収益(OBC)
                air_32117=air_32117+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32119 in row[ACCOUNT_POS]):#航空輸出収益  その他
                air_32119=air_32119+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32121 in row[ACCOUNT_POS]):#航空輸出費用  混載支払
                air_32121=air_32121+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32123 in row[ACCOUNT_POS]):#航空輸出費用  入出庫費
                air_32123=air_32123+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32124 in row[ACCOUNT_POS]):#航空輸出費用  委託運送費
                air_32124=air_32124+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32125 in row[ACCOUNT_POS]):#航空輸出費用  情報処理費  要確認
                air_32125=air_32125+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32126 in row[ACCOUNT_POS]):#航空輸出費用  通関費
                air_32126=air_32126+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32127 in row[ACCOUNT_POS]):#航空輸出費用  クーリエ費用(OBC)
                air_32127=air_32127+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_32128 in row[ACCOUNT_POS]):#航空輸出費用  その他"
                air_32128=air_32128+ int(row[TOGETSU_POS]) #当月金額

    air_32111=Ushiro_five_handred_rounding_off(air_32111)
    air_32112=Ushiro_five_handred_rounding_off(air_32112)
    air_32113=Ushiro_five_handred_rounding_off(air_32113)
    air_32114=Ushiro_five_handred_rounding_off(air_32114)
    air_32116=Ushiro_five_handred_rounding_off(air_32116)
    air_32117=Ushiro_five_handred_rounding_off(air_32117)
    air_32119=Ushiro_five_handred_rounding_off(air_32119)
    incom_sum = air_32111+air_32112+air_32113+air_32114+air_32116+air_32117+air_32119#収入合計:転記不要
    air_32121=Ushiro_five_handred_rounding_off(air_32121)
    air_32123=Ushiro_five_handred_rounding_off(air_32123)
    air_32124=Ushiro_five_handred_rounding_off(air_32124)
    air_32125=Ushiro_five_handred_rounding_off(air_32125)
    air_32126=Ushiro_five_handred_rounding_off(air_32126)
    air_32127=Ushiro_five_handred_rounding_off(air_32127)
    air_32128=Ushiro_five_handred_rounding_off(air_32128)

    outcom_sum = air_32121+air_32123+air_32124+air_32125+air_32127+air_32128 #支出合計：転記不要

    prft = incom_sum - outcom_sum #差益：転記不要

    if deb_flg ==1:
        print("合計：", air_32111)
        print("合計：", air_32112)
        print("合計：", air_32113)
        print("合計：", air_32114)
        print("合計：", air_32116)
        print("合計：", air_32117)
        print("合計：", air_32119)
        print("収入合計：", incom_sum)#収入計
        print("合計：", air_32121)
        print("合計：", air_32123)
        print("合計：", air_32124)
        print("合計：", air_32125)
        print("合計：", air_32127)
        print("合計：", air_32128)
        print("支出合計：", outcom_sum)#支出計
        print("差益：", prft)#差益
    #------------------------------------------------------------------------
    wb2 = openpyxl.load_workbook(OUTPUT_FILE)
    print(wb2.get_sheet_names())
    print("*------2019年度(第31期)営業収支実績.xlsx------*")
    #sheet2=wb2.get_sheet_by_name("4月")
    sheet2 = wb2.get_sheet_by_name(AIR_TSUKI)

    sheet2[KONSAI]=air_32111
    sheet2[TESURY]=air_32112
    sheet2[NYUSYU]=air_32113
    sheet2[UNSORO]=air_32114
    sheet2[TSUKAN]=air_32116
    sheet2[OBCSYU]=air_32117
    sheet2[SONOTA]=air_32119
    sheet2[KONGEN]=air_32121
    sheet2[NYUSYH]=air_32123
    sheet2[UNSOHI]=air_32124
    sheet2[TSUKAH]=air_32126
    sheet2[OBCHIY]=air_32127
    sheet2[SONOTH]=air_32128+air_32125
    #営業----------------------------------------------------*

    wb2.save(OUTPUT_FILE) #ファイルへの書き込み
#------------------------------------------------------#
#成田・羽田・関空
#------------------------------------------------------#
def exec_eigyosho(cnt):#成田・羽田・関空
    if cnt==0:#成田
        FILE_eigyosho = AIR_NARITA_F
    elif cnt==1:#羽田
        FILE_eigyosho = AIR_HANED_F
    elif cnt==2:#関空
        FILE_eigyosho = AIR_KANKU_F

    AC_AIR_322111 = "322111" #混載手数料その１
    AC_AIR_322112 = "322112" #混載手数料その２
    AC_AIR_322114 = "322114" #混載手数料その３
    AC_AIR_322114 = "322114" #混載手数料その４
    AC_AIR_322118 = "322118" #着払い運賃
    AC_AIR_322130 = "322130" #入出庫料
    AC_AIR_322140 = "322140" #運送料
    AC_AIR_322160 = "322160" #通関料
    AC_AIR_322190 = "322190" #その他
    AC_AIR_322230 = "322230" #出庫費
    AC_AIR_322241 = "322241" #運送費
    AC_AIR_322242 = "322242" #運送傭車費
    AC_AIR_322250 = "322250" #情報処理費
    AC_AIR_322260 = "322260" #通関費
    AC_AIR_322289 = "322289" #その他

    #値の初期化
    air_322111 = 0 #混載手数料その１
    air_322112 = 0 #混載手数料その２
    air_322114 = 0 #混載手数料その３
    air_322118 = 0 #着払い運賃
    air_322130 = 0 #入出庫料
    air_322140 = 0 #運送料
    air_322160 = 0 #通関料
    air_322190 = 0 #その他

    air_322230 = 0 #出庫費
    air_322241 = 0 #運送費
    air_322242 = 0 #運送傭車費
    air_322250 = 0 #情報処理費
    air_322260 = 0 #通関費
    air_322289 = 0 #その他

    ACCOUNT_POS=0 #勘定科目の列番目
    TOGETSU_POS=4#当月金額の列番目

    #転送先セル-------------------------------------------------*
    #収入
    if cnt==0:#成田
        CC_FEE = "K26" #ＣＣＦＥＥ
        COMITT = "K28" #コミッション
        TESURY = "K29" #手数料
        NYUSYU = "K30" #入出庫
        UNSORO = "K31" #運送料
        TSUKAN = "K32" #通関料
        SONOTA = "K34" #その他
       #費用
        NYUSYH = "K36" #入出庫費
        UNSOHI = "K37" #運送費
        TSUKAH = "K38" #通関費
        SONOTH = "K40" #その他費用
    elif cnt==1:#羽田
        CC_FEE = "P26"  # ＣＣＦＥＥ
        COMITT = "P28"  # コミッション
        TESURY = "P29"  # 手数料
        NYUSYU = "P30"  # 入出庫
        UNSORO = "P31"  # 運送料
        TSUKAN = "P32"  # 通関料
        SONOTA = "P34"  # その他
        # 費用
        NYUSYH = "P36"  # 入出庫費
        UNSOHI = "P37"  # 運送費
        TSUKAH = "P38"  # 通関費
        SONOTH = "P40"  # その他費用
    elif cnt == 2:# 関空
        CC_FEE = "U26"  # ＣＣＦＥＥ
        COMITT = "U28"  # コミッション
        TESURY = "U29"  # 手数料
        NYUSYU = "U30"  # 入出庫
        UNSORO = "U31"  # 運送料
        TSUKAN = "U32"  # 通関料
        SONOTA = "U34"  # その他
        # 費用
        NYUSYH = "U36"  # 入出庫費
        UNSOHI = "U37"  # 運送費
        TSUKAH = "U38"  # 通関費
        SONOTH = "U40"  # その他費用
    #----------------------------------------------------*
    cntr=0
    with open(FILE_eigyosho) as f:
        #next(f)#１行目がブランクだとエラーになるので苦肉の策（２行目から読む）
        reader = csv.reader(f)
        for row in reader:
            cntr=cntr+1
            if(AC_AIR_322111 in row[ACCOUNT_POS]):#航空輸出収益  混載その１
                air_322111=air_322111+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322112 in row[ACCOUNT_POS]):#航空輸出収益 混載その２
                air_322112=air_322112+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322114 in row[ACCOUNT_POS]):#航空輸出収益  混載その３
                air_322114=air_322114+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322118 in row[ACCOUNT_POS]):#航空輸出収益  着払い運賃
                air_322118=air_322118+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322130 in row[ACCOUNT_POS]):#航空輸出収益  入出庫料
                air_322130=air_322130+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322140 in row[ACCOUNT_POS]):#航空輸出収益  運送料
                air_322140=air_322140+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322160 in row[ACCOUNT_POS]):#航空輸出収益  通関料
                air_322160=air_322160+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322190 in row[ACCOUNT_POS]):#航空輸出収益  その他
                air_322190=air_322190+ int(row[TOGETSU_POS]) #当月金額
         #費用
            elif(AC_AIR_322230 in row[ACCOUNT_POS]):#航空輸出収益  出庫費
                air_322230=air_322230+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322241 in row[ACCOUNT_POS]):#航空輸出収益  運送費
                air_322241=air_322241+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322242 in row[ACCOUNT_POS]):#航空輸出収益  傭車費
                air_322242=air_322242+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322250 in row[ACCOUNT_POS]):#航空輸出収益  情報処理
                air_322250=air_322250+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322260 in row[ACCOUNT_POS]):#航空輸出収益  通関費
                air_322260=air_322260+ int(row[TOGETSU_POS]) #当月金額
            elif(AC_AIR_322289 in row[ACCOUNT_POS]):#航空輸出収益  その他
                air_322289=air_322289+ int(row[TOGETSU_POS]) #当月金額

    air_322111=Ushiro_five_handred_rounding_off(air_322111)
    air_322112=Ushiro_five_handred_rounding_off(air_322112)
    air_322130=Ushiro_five_handred_rounding_off(air_322130)
    air_322140=Ushiro_five_handred_rounding_off(air_322140)
    air_322160=Ushiro_five_handred_rounding_off(air_322160)
    air_322190=Ushiro_five_handred_rounding_off(air_322190)

    air_322230=Ushiro_five_handred_rounding_off(air_322230)
    air_322260=Ushiro_five_handred_rounding_off(air_322260)

    ccfee=Ushiro_five_handred_rounding_off(air_322114 + air_322118)
    itk_unso=Ushiro_five_handred_rounding_off((air_322241+air_322242))
    sonota=Ushiro_five_handred_rounding_off(air_322250 + air_322289)

    #------------------------------------------------------------------------
    wb2 = openpyxl.load_workbook(OUTPUT_FILE)
    print("*------2019年度(第31期)営業収支実績.xlsx------*")
    sheet2=wb2.get_sheet_by_name(AIR_TSUKI)

    sheet2[CC_FEE] = ccfee
    sheet2[COMITT]=air_322112
    sheet2[TESURY]=air_322111
    sheet2[NYUSYU]=air_322130
    sheet2[UNSORO]=air_322140
    sheet2[TSUKAN]=air_322160
    sheet2[SONOTA]=air_322190
#費用
    sheet2[NYUSYH]=air_322230
    sheet2[UNSOHI]=itk_unso
    sheet2[TSUKAH]=air_322260
    sheet2[SONOTH]=sonota
    wb2.save(OUTPUT_FILE) #ファイルへの書き込み
#------------------------------------------------------#
#管理費配賦・業績評価
#------------------------------------------------------#
def kanrihi_gyoseki_hyoka():
    #管理費配賦
    Haifu_F=AIR_HAIFU_F
    wb_knr = openpyxl.load_workbook(Haifu_F, data_only=True)
    #業績評価
    Haifu_G=AIR_GYOSEKI_F
    wb_gyoseki = openpyxl.load_workbook(Haifu_G, data_only=True)
    wb2 = openpyxl.load_workbook(OUTPUT_FILE)
    print("*------営業収支実績------*")
    sheet2=wb2.get_sheet_by_name(AIR_TSUKI)

    print("*------配賦セット！------*")
    sheet_wk = wb_knr.get_sheet_by_name(AIR_TSUKI)  # 管理費配賦表

    sheet2['F48'] = sheet_wk['G8'].value#関東営業
    sheet2['P48'] = sheet_wk['G9'].value#羽田
    sheet2['K48'] = sheet_wk['G10'].value#成田
    sheet2['Z48'] = sheet_wk['G12'].value#関西営業
    sheet2['U48'] = sheet_wk['G13'].value#関空

    print("*------業績セット！------*")
    sheet_wk = wb_gyoseki.get_sheet_by_name("Sheet1")  #業績表
    sheet2['F46'] = round(sheet_wk['P53'].value/1000)#関東営業
    sheet2['K46'] = round(sheet_wk['Z53'].value/1000)#成田
    sheet2['P46'] = round(sheet_wk['AE53'].value/1000)#羽田
    sheet2['Z46'] = round(sheet_wk['AJ53'].value/1000)#関西営業
    sheet2['U46'] = round(sheet_wk['AO53'].value/1000)#関空

    wb2.save(OUTPUT_FILE) #ファイルへの書き込み
#--------------------------------------------------------------------#
#  10000で割った値の後ろ3桁が500のとき、四捨五入が効かないので強引に
# 501にして、四捨五入を効かす関数
# 例：☓2466500→2466 　○2466500→2467
#--------------------------------------------------------------------#
def Ushiro_five_handred_rounding_off(val):
    #val = 2466500
    m_val = str(val)#文字列に変換
    last_3_keta = m_val[-3:]#後ろ3桁を抽出
    if last_3_keta == '500':
        val = val + 1  #501にする
    val = round(val / 1000)#すると2467になる（例）これをやらないと2466になってしまう
    return val

#------------------------------------------------------#
#メインルーチン
#------------------------------------------------------#
if __name__ == '__main__':
    pyautogui.alert('ケイヒン航空PL転記システムをスタートします')
    initial()#初期処理
    for cnt in range(2):
        exec_sales(cnt)#関東・関西営業PLセット
    for cnt in range(3):
        exec_eigyosho(cnt)#成田・羽田・関空PLセット
    kanrihi_gyoseki_hyoka()#管理費配賦・業績評価
    pyautogui.alert('システムが終了しました')
