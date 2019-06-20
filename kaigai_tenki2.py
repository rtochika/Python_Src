import time
from selenium import webdriver
import pyautogui
import openpyxl
import csv

deb_flg =0# 1:ON 0:OFF
#【機能】---------------------------------------------------*
#海外現地法人の収支実績を転記する
# インプット・アウトプット対比表 → http://bit.ly/2XDFG9J
#-----------------------------------------------------------*
#-----------------------------------------------------------------------------------------#
pyautogui.alert("海外現法PLデータ転記を開始します","Excel_Move_Dta")#プログラム開始確認
#-----------------------------------------------------------------------------------------#
#ファイル名処理-----------------------------------------------------------*

INPUT_CSV='kaig_ten_input_file.txt'
#INPUT_CSV=r'C:\Users\86001\PycharmProjects\HelloTensorFlow\kaig_ten_input_file.csv'
#INPUT_CSV='kaig_ten_input_file.csv'

cntr=0
with open(INPUT_CSV) as f:
    reader = csv.reader(f)
    for row in reader:
        if row[0] !="#":
            cntr=cntr+1
            if cntr==1:
                OUTPUT_FILE = row[0]  # 例：海外18012.xlsx
            elif cntr == 2:
                SNG_IN_F = row[0]  # KMTS2018.12.xlsx
            elif cntr == 3:
                HNK_IN_F = row[0]  # KMTH2018.12.xlsx
            elif cntr == 4:
                SHA_IN_F = row[0]  # KMT上海2018.12.xlsx
            elif cntr == 5:
                PHI_IN_F = row[0]  # KEFCI2018.12 .xlsx
            elif cntr == 6:
                TWN_IN_F = row[0]  # KMTT2018.12.xlsx
            elif cntr == 7:
                VTN_IN_F = row[0]  # KMTV2018.12.xlsx

if deb_flg == 1:
    print(OUTPUT_FILE)
    print(SNG_IN_F)
    print(HNK_IN_F)
    print(SHA_IN_F)
    print(PHI_IN_F)
    print(TWN_IN_F)
    print(VTN_IN_F)

#ファイル名
#------------------------------------------------
#OUTPUT_FILE = "海外18012.xlsx"
#SNG_IN_F = "【C】KMTS2018.12.xlsx"
#HNK_IN_F = "【C】KMTH2018.12.xlsx"
#SHA_IN_F = "【C】KMT上海2018.12.xlsx"
#PHI_IN_F = "【C】KEFCI2018.12 .xlsx"
#TWN_IN_F = "【C】KMTT2018.12.xlsx"
#VTN_IN_F = "【C】KMTV2018.12.xlsx"
#------------------------------------------------

from selenium.webdriver.remote.webelement import WebElement

#閲覧文書終始実績Excelデータ操作
#wb2 = openpyxl.load_workbook("海外18012.xlsx")
wb2 = openpyxl.load_workbook(OUTPUT_FILE.strip())

#print(wb2.get_sheet_names())
print("*------シンガポール開始！------*")
sheet2=wb2.get_sheet_by_name("S")#シンガポールPL（アウトプット）

#シンガポールExcelデータ操作
#wb1 = openpyxl.load_workbook("【C】KMTS2018.12.xlsx", data_only=True)
wb1 = openpyxl.load_workbook(SNG_IN_F.strip(), data_only=True)

#print(wb1.get_sheet_names())
sheet1=wb1.get_sheet_by_name("PL")#シンガポールPL（インプット）

#シンガポール------------------------------------------*
for cnt in range(9,58+1):
    sheet2['I'+str(cnt)]=sheet1['I'+str(cnt)].value
    sheet2['K'+str(cnt)]=sheet1['J'+str(cnt)].value
    sheet2['M'+str(cnt)]=sheet1['K'+str(cnt)].value
    sheet2['O'+str(cnt)]=sheet1['L'+str(cnt)].value
    sheet2['Q'+str(cnt)]=sheet1['M'+str(cnt)].value
    sheet2['S'+str(cnt)]=sheet1['N'+str(cnt)].value

#香港----------------------------------------------------*
sheet2=wb2.get_sheet_by_name("HK")#香港（アウトプット）
#wb_hk = openpyxl.load_workbook("【C】KMTH2018.12.xlsx", data_only=True)
wb_hk = openpyxl.load_workbook(HNK_IN_F.strip(), data_only=True)

#print(wb_hk.get_sheet_names())
print("*------香港開始！------*")
sheet3=wb_hk.get_sheet_by_name("PL")#香港PL（インプット）
for cnt in range(9,60+1):
    if(cnt<=51):
        sheet2['I'+str(cnt)]=sheet3['I'+str(cnt)].value
        sheet2['K'+str(cnt)]=sheet3['J'+str(cnt)].value
        sheet2['M'+str(cnt)]=sheet3['K'+str(cnt)].value
        sheet2['O'+str(cnt)]=sheet3['L'+str(cnt)].value
        sheet2['Q'+str(cnt)]=sheet3['M'+str(cnt)].value
        sheet2['S'+str(cnt)]=sheet3['N'+str(cnt)].value
    #elif (cnt >= 53):
    elif(cnt>53):
        sheet2['I' + str(cnt-2)] = sheet3['I' + str(cnt)].value
        sheet2['K' + str(cnt-2)] = sheet3['J' + str(cnt)].value
        sheet2['M' + str(cnt-2)] = sheet3['K' + str(cnt)].value
        sheet2['O' + str(cnt-2)] = sheet3['L' + str(cnt)].value
        sheet2['Q' + str(cnt-2)] = sheet3['M' + str(cnt)].value
        sheet2['S' + str(cnt-2)] = sheet3['N' + str(cnt)].value

#上海----------------------------------------------------*
sheet2=wb2.get_sheet_by_name("SH")#上海（アウトプット）
#wb_sh = openpyxl.load_workbook("【C】KMT上海2018.12.xlsx", data_only=True)
wb_sh = openpyxl.load_workbook(SHA_IN_F.strip(), data_only=True)

#print(wb_sh.get_sheet_names())
print("*------上海開始！------*")
sheet4=wb_sh.get_sheet_by_name("3.PL")#上海PL（インプット）
for cnt in range(9,48+1):
    #if(cnt<=28):
    if (cnt <= 33):
        sheet2['I'+str(cnt)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt)]=sheet4['P'+str(cnt)].value
    # elif(cnt>=34 and cnt <=41):
    elif (cnt >= 34 and cnt <= 42):
        sheet2['I'+str(cnt+10)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt+10)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt+10)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt+10)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt+10)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt+10)]=sheet4['P'+str(cnt)].value
    elif(cnt==43):
        sheet2['I'+str(cnt+10)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt+10)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt+10)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt+10)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt+10)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt+10)]=sheet4['P'+str(cnt)].value
    elif (cnt >=46):
        sheet2['I'+str(cnt+10)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt+10)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt+10)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt+10)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt+10)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt+10)]=sheet4['P'+str(cnt)].value
#print("I52",sheet2['I52'].value)

#フィリピン----------------------------------------------------*
sheet2=wb2.get_sheet_by_name("Ph")#フィリピン（アウトプット）
#wb_pi = openpyxl.load_workbook("【C】KEFCI2018.12 .xlsx", data_only=True)
wb_pi = openpyxl.load_workbook(PHI_IN_F.strip(), data_only=True)
#print(wb_pi.get_sheet_names())
print("*------フィリピン開始！------*")
sheet5=wb_pi.get_sheet_by_name("PL")#フィリピンPL（インプット）
for cnt in range(9,63+1):
    if(cnt<=23):
        if(deb_flg):
            print("output cnt = ",cnt)
        sheet2['I'+str(cnt)]=sheet5['I'+str(cnt)].value
        sheet2['K'+str(cnt)]=sheet5['J'+str(cnt)].value
        sheet2['M'+str(cnt)]=sheet5['K'+str(cnt)].value
        sheet2['O'+str(cnt)]=sheet5['L'+str(cnt)].value
        sheet2['Q'+str(cnt)]=sheet5['M'+str(cnt)].value
        sheet2['S'+str(cnt)]=sheet5['N'+str(cnt)].value
    elif(cnt>=24 and cnt<=28):
        if(deb_flg):
            print("output cnt = ",cnt)
        sheet2['I'+str(cnt)]=sheet5['I'+str(cnt)].value + sheet5['I'+str(cnt+5)].value
        sheet2['K'+str(cnt)]=sheet5['J'+str(cnt)].value + sheet5['J'+str(cnt+5)].value
        sheet2['M'+str(cnt)]=sheet5['K'+str(cnt)].value + sheet5['K'+str(cnt+5)].value
        sheet2['O'+str(cnt)]=sheet5['L'+str(cnt)].value + sheet5['L'+str(cnt+5)].value
        sheet2['Q'+str(cnt)]=sheet5['M'+str(cnt)].value + sheet5['M'+str(cnt+5)].value
        sheet2['S'+str(cnt)]=sheet5['N'+str(cnt)].value + sheet5['N'+str(cnt+5)].value
    elif(cnt==29):
        cnt=33
    elif(cnt>=34 and cnt<=38):
        if(deb_flg):
            print("output cnt = ",cnt+3)
        sheet2['I'+str(cnt+3)]=sheet5['I'+str(cnt)].value
        sheet2['K'+str(cnt+3)]=sheet5['J'+str(cnt)].value
        sheet2['M'+str(cnt+3)]=sheet5['K'+str(cnt)].value
        sheet2['O'+str(cnt+3)]=sheet5['L'+str(cnt)].value
        sheet2['Q'+str(cnt+3)]=sheet5['M'+str(cnt)].value
        sheet2['S'+str(cnt+3)]=sheet5['N'+str(cnt)].value
    elif(cnt>=39 and cnt<=41):
        if(deb_flg):
            print("output cnt = ",cnt-10)
        sheet2['I'+str(cnt-10)]=sheet5['I'+str(cnt)].value
        sheet2['K'+str(cnt-10)]=sheet5['J'+str(cnt)].value
        sheet2['M'+str(cnt-10)]=sheet5['K'+str(cnt)].value
        sheet2['O'+str(cnt-10)]=sheet5['L'+str(cnt)].value
        sheet2['Q'+str(cnt-10)]=sheet5['M'+str(cnt)].value
        sheet2['S'+str(cnt-10)]=sheet5['N'+str(cnt)].value
        if(cnt==41):
            if (deb_flg):
                print("output cnt = ", cnt)
            sheet2['I' + str(34)] = sheet5['I' + str(cnt)].value
            sheet2['K' + str(34)] = sheet5['J' + str(cnt)].value
            sheet2['M' + str(34)] = sheet5['K' + str(cnt)].value
            sheet2['O' + str(34)] = sheet5['L' + str(cnt)].value
            sheet2['Q' + str(34)] = sheet5['M' + str(cnt)].value
            sheet2['S' + str(34)] = sheet5['N' + str(cnt)].value
    elif(cnt>=42 and cnt<=43):
        if(deb_flg):
            print("output cnt = ",cnt)
        sheet2['I' + str(cnt-7)] = sheet5['I' + str(cnt)].value
        sheet2['K' + str(cnt-7)] = sheet5['J' + str(cnt)].value
        sheet2['M' + str(cnt-7)] = sheet5['K' + str(cnt)].value
        sheet2['O' + str(cnt-7)] = sheet5['L' + str(cnt)].value
        sheet2['Q' + str(cnt-7)] = sheet5['M' + str(cnt)].value
        sheet2['S' + str(cnt-7)] = sheet5['N' + str(cnt)].value
        if (cnt == 43):
            cnt=48
    elif(cnt>=49 and cnt<=51):
        if(deb_flg):
            print("output cnt = ",cnt)
        sheet2['I' + str(cnt - 7)] = sheet5['I' + str(cnt)].value
        sheet2['K' + str(cnt - 7)] = sheet5['J' + str(cnt)].value
        sheet2['M' + str(cnt - 7)] = sheet5['K' + str(cnt)].value
        sheet2['O' + str(cnt - 7)] = sheet5['L' + str(cnt)].value
        sheet2['Q' + str(cnt - 7)] = sheet5['M' + str(cnt)].value
        sheet2['S' + str(cnt - 7)] = sheet5['N' + str(cnt)].value
        if(cnt==51):
            if (deb_flg):
                print("output cnt = ", cnt)
            sheet2['I' + str(cnt - 4)] = sheet5['I' + str(cnt)].value
            sheet2['K' + str(cnt - 4)] = sheet5['J' + str(cnt)].value
            sheet2['M' + str(cnt - 4)] = sheet5['K' + str(cnt)].value
            sheet2['O' + str(cnt - 4)] = sheet5['L' + str(cnt)].value
            sheet2['Q' + str(cnt - 4)] = sheet5['M' + str(cnt)].value
            sheet2['S' + str(cnt - 4)] = sheet5['N' + str(cnt)].value
    elif(cnt>=52):
        if(deb_flg):
            print("output cnt = ", cnt)
        sheet2['I' + str(cnt - 4)] = sheet5['I' + str(cnt)].value
        sheet2['K' + str(cnt - 4)] = sheet5['J' + str(cnt)].value
        sheet2['M' + str(cnt - 4)] = sheet5['K' + str(cnt)].value
        sheet2['O' + str(cnt - 4)] = sheet5['L' + str(cnt)].value
        sheet2['Q' + str(cnt - 4)] = sheet5['M' + str(cnt)].value
        sheet2['S' + str(cnt - 4)] = sheet5['N' + str(cnt)].value
#台湾----------------------------------------------------*
sheet2=wb2.get_sheet_by_name("T")#台湾（アウトプット）
#wb_tw = openpyxl.load_workbook("【C】KMTT2018.12.xlsx", data_only=True)
wb_tw = openpyxl.load_workbook(TWN_IN_F.strip(), data_only=True)
#print(wb_tw.get_sheet_names())
print("*------台湾開始！------*")
sheet6=wb_tw.get_sheet_by_name("PL")#台湾PL（インプット）
for cnt in range(9,58+1):
    sheet2['I'+str(cnt)]=sheet6['I'+str(cnt)].value
    sheet2['K'+str(cnt)]=sheet6['J'+str(cnt)].value
    sheet2['M'+str(cnt)]=sheet6['K'+str(cnt)].value
    sheet2['O'+str(cnt)]=sheet6['L'+str(cnt)].value
    sheet2['Q'+str(cnt)]=sheet6['M'+str(cnt)].value
    sheet2['S'+str(cnt)]=sheet6['N'+str(cnt)].value

#ベトナム----------------------------------------------------*
sheet2=wb2.get_sheet_by_name("V")#ベトナム（アウトプット）
#wb_tv = openpyxl.load_workbook("【C】KMTV2018.12.xlsx", data_only=True)
wb_tv = openpyxl.load_workbook(VTN_IN_F.strip(), data_only=True)
#print(wb_tv.get_sheet_names())
sheet7=wb_tv.get_sheet_by_name("PL")#ベトナムPL（インプット）
print("*------ベトナム開始！------*")
#In-Data  B C D E F G列  #Out-Data I K M O Q S列
#http://bit.ly/2IBmcPP
#----------------
sheet2['I9'] =  sheet7['B26'].value
sheet2['K9'] =  sheet7['C26'].value
sheet2['M9'] =  sheet7['D26'].value
sheet2['O9'] =  sheet7['E26'].value
sheet2['Q9'] =  sheet7['F26'].value
sheet2['S9'] =  sheet7['G26'].value
#----------------
sheet2['I10'] = sheet7['B27'].value
sheet2['K10'] = sheet7['C27'].value
sheet2['M10'] = sheet7['D27'].value
sheet2['O10'] = sheet7['E27'].value
sheet2['Q10'] = sheet7['F27'].value
sheet2['S10'] = sheet7['G27'].value
#----------------
sheet2['I11'] = sheet7['B28'].value
sheet2['K11'] = sheet7['C28'].value
sheet2['M11'] = sheet7['D28'].value
sheet2['O11'] = sheet7['E28'].value
sheet2['Q11'] = sheet7['F28'].value
sheet2['S11'] = sheet7['G28'].value
#----------------
sheet2['I12'] =  0
sheet2['K12'] =  0
sheet2['M12'] =  0
sheet2['O12'] =  0
sheet2['Q12'] =  0
sheet2['S12'] =  0
#----------------
sheet2['I13'] = sheet2['I11'].value - sheet2['I12'].value
sheet2['K13'] = sheet2['K11'].value - sheet2['K12'].value
sheet2['M13'] = sheet2['M11'].value - sheet2['M12'].value
sheet2['O13'] = sheet2['O11'].value - sheet2['O12'].value
sheet2['Q13'] = sheet2['Q11'].value - sheet2['Q12'].value
sheet2['S13'] = sheet2['S11'].value - sheet2['S12'].value
#----------------
sheet2['I14'] = sheet7['B30'].value
sheet2['K14'] = sheet7['C30'].value
sheet2['M14'] = sheet7['D30'].value
sheet2['O14'] = sheet7['E30'].value
sheet2['Q14'] = sheet7['F30'].value
sheet2['S14'] = sheet7['G30'].value
#----------------
sheet2['I15'] = sheet7['B31'].value
sheet2['K15'] = sheet7['C31'].value
sheet2['M15'] = sheet7['D31'].value
sheet2['O15'] = sheet7['E31'].value
sheet2['Q15'] = sheet7['F31'].value
sheet2['S15'] = sheet7['G31'].value
#----------------
sheet2['I16'] = sheet7['B32'].value
sheet2['K16'] = sheet7['C32'].value
sheet2['M16'] = sheet7['D32'].value
sheet2['O16'] = sheet7['E32'].value
sheet2['Q16'] = sheet7['F32'].value
sheet2['S16'] = sheet7['G32'].value
#----------------
sheet2['I17'] = 0
sheet2['K17'] = 0
sheet2['M17'] = 0
sheet2['O17'] = 0
sheet2['Q17'] = 0
sheet2['S17'] = 0
#----------------
sheet2['I18'] = sheet2['I16'].value - sheet2['I17'].value
sheet2['K18'] = sheet2['K16'].value - sheet2['K17'].value
sheet2['M18'] = sheet2['M16'].value - sheet2['M17'].value
sheet2['O18'] = sheet2['O16'].value - sheet2['O17'].value
sheet2['Q18'] = sheet2['Q16'].value - sheet2['Q17'].value
sheet2['S18'] = sheet2['S16'].value - sheet2['S17'].value
#----------------
sheet2['I19'] = sheet2['I9'].value + sheet2['I14'].value
sheet2['K19'] = sheet2['K9'].value + sheet2['K14'].value
sheet2['M19'] = sheet2['M9'].value + sheet2['M14'].value
sheet2['O19'] = sheet2['O9'].value + sheet2['O14'].value
sheet2['Q19'] = sheet2['Q9'].value + sheet2['Q14'].value
sheet2['S19'] = sheet2['S9'].value + sheet2['S14'].value
#----------------
sheet2['I20'] = sheet2['I10'].value + sheet2['I15'].value
sheet2['K20'] = sheet2['K10'].value + sheet2['K15'].value
sheet2['M20'] = sheet2['M10'].value + sheet2['M15'].value
sheet2['O20'] = sheet2['O10'].value + sheet2['O15'].value
sheet2['Q20'] = sheet2['Q10'].value + sheet2['Q15'].value
sheet2['S20'] = sheet2['S10'].value + sheet2['S15'].value
#----------------
sheet2['I21'] = sheet2['I11'].value + sheet2['I16'].value
sheet2['K21'] = sheet2['K11'].value + sheet2['K16'].value
sheet2['M21'] = sheet2['M11'].value + sheet2['M16'].value
sheet2['O21'] = sheet2['O11'].value + sheet2['O16'].value
sheet2['Q21'] = sheet2['Q11'].value + sheet2['Q16'].value
sheet2['S21'] = sheet2['S11'].value + sheet2['S16'].value
#----------------
sheet2['I22'] = sheet2['I12'].value + sheet2['I17'].value
sheet2['K22'] = sheet2['K12'].value + sheet2['K17'].value
sheet2['M22'] = sheet2['M12'].value + sheet2['M17'].value
sheet2['O22'] = sheet2['O12'].value + sheet2['O17'].value
sheet2['Q22'] = sheet2['Q12'].value + sheet2['Q17'].value
sheet2['S22'] = sheet2['S12'].value + sheet2['S17'].value
#----------------
sheet2['I23'] = sheet2['I13'].value + sheet2['I18'].value
sheet2['K23'] = sheet2['K13'].value + sheet2['K18'].value
sheet2['M23'] = sheet2['M13'].value + sheet2['M18'].value
sheet2['O23'] = sheet2['O13'].value + sheet2['O18'].value
sheet2['Q23'] = sheet2['Q13'].value + sheet2['Q18'].value
sheet2['S23'] = sheet2['S13'].value + sheet2['S18'].value
#----------------
sheet2['I24'] = sheet7['B9'].value
sheet2['K24'] = sheet7['C9'].value
sheet2['M24'] = sheet7['D9'].value
sheet2['O24'] = sheet7['E9'].value
sheet2['Q24'] = sheet7['F9'].value
sheet2['S24'] = sheet7['G9'].value
#----------------
sheet2['I25'] = sheet7['B10'].value
sheet2['K25'] = sheet7['C10'].value
sheet2['M25'] = sheet7['D10'].value
sheet2['O25'] = sheet7['E10'].value
sheet2['Q25'] = sheet7['F10'].value
sheet2['S25'] = sheet7['G10'].value
#----------------
sheet2['I26'] = sheet7['B11'].value
sheet2['K26'] = sheet7['C11'].value
sheet2['M26'] = sheet7['D11'].value
sheet2['O26'] = sheet7['E11'].value
sheet2['Q26'] = sheet7['F11'].value
sheet2['S26'] = sheet7['G11'].value
#----------------
sheet2['I27'] = 0
sheet2['K27'] = 0
sheet2['M27'] = 0
sheet2['O27'] = 0
sheet2['Q27'] = 0
sheet2['S27'] = 0
#----------------
sheet2['I28'] = sheet2['I26'].value
sheet2['K28'] = sheet2['K26'].value
sheet2['M28'] = sheet2['M26'].value
sheet2['O28'] = sheet2['O26'].value
sheet2['Q28'] = sheet2['Q26'].value
sheet2['S28'] = sheet2['S26'].value
#----------------
sheet2['I29'] = sheet7['B13'].value
sheet2['K29'] = sheet7['C13'].value
sheet2['M29'] = sheet7['D13'].value
sheet2['O29'] = sheet7['E13'].value
sheet2['Q29'] = sheet7['F13'].value
sheet2['S29'] = sheet7['G13'].value
#----------------
sheet2['I30'] = sheet7['B14'].value
sheet2['K30'] = sheet7['C14'].value
sheet2['M30'] = sheet7['D14'].value
sheet2['O30'] = sheet7['E14'].value
sheet2['Q30'] = sheet7['F14'].value
sheet2['S30'] = sheet7['G14'].value
#----------------
sheet2['I31'] = sheet7['B15'].value
sheet2['K31'] = sheet7['C15'].value
sheet2['M31'] = sheet7['D15'].value
sheet2['O31'] = sheet7['E15'].value
sheet2['Q31'] = sheet7['F15'].value
sheet2['S31'] = sheet7['G15'].value
#----------------
sheet2['I32'] = 0
sheet2['K32'] = 0
sheet2['M32'] = 0
sheet2['O32'] = 0
sheet2['Q32'] = 0
sheet2['S32'] = 0
#----------------
sheet2['I33'] = sheet2['I31'].value - sheet2['I32'].value
sheet2['K33'] = sheet2['K31'].value - sheet2['K32'].value
sheet2['M33'] = sheet2['M31'].value - sheet2['M32'].value
sheet2['O33'] = sheet2['O31'].value - sheet2['O32'].value
sheet2['Q33'] = sheet2['Q31'].value - sheet2['Q32'].value
sheet2['S33'] = sheet2['S31'].value - sheet2['S32'].value
#----------------
sheet2['I34'] = sheet7['B17'].value
sheet2['K34'] = sheet7['C17'].value
sheet2['M34'] = sheet7['D17'].value
sheet2['O34'] = sheet7['E17'].value
sheet2['Q34'] = sheet7['F17'].value
sheet2['S34'] = sheet7['G17'].value
#----------------
sheet2['I35'] = sheet7['B18'].value
sheet2['K35'] = sheet7['C18'].value
sheet2['M35'] = sheet7['D18'].value
sheet2['O35'] = sheet7['E18'].value
sheet2['Q35'] = sheet7['F18'].value
sheet2['S35'] = sheet7['G18'].value
#----------------
sheet2['I36'] = sheet7['B19'].value
sheet2['K36'] = sheet7['C19'].value
sheet2['M36'] = sheet7['D19'].value
sheet2['O36'] = sheet7['E19'].value
sheet2['Q36'] = sheet7['F19'].value
sheet2['S36'] = sheet7['G19'].value
#----------------
sheet2['I37'] = 0
sheet2['K37'] = 0
sheet2['M37'] = 0
sheet2['O37'] = 0
sheet2['Q37'] = 0
sheet2['S37'] = 0
#----------------
sheet2['I38'] = 0
sheet2['K38'] = 0
sheet2['M38'] = 0
sheet2['O38'] = 0
sheet2['Q38'] = 0
sheet2['S38'] = 0
#----------------
sheet2['I39'] = sheet2['I36'].value
sheet2['K39'] = sheet2['K36'].value
sheet2['M39'] = sheet2['M36'].value
sheet2['O39'] = sheet2['O36'].value
sheet2['Q39'] = sheet2['Q36'].value
sheet2['S39'] = sheet2['S36'].value
#----------------
sheet2['I40'] = 0
sheet2['K40'] = 0
sheet2['M40'] = 0
sheet2['O40'] = 0
sheet2['Q40'] = 0
sheet2['S40'] = 0
#----------------
sheet2['I41'] = sheet2['I39'].value
sheet2['K41'] = sheet2['K39'].value
sheet2['M41'] = sheet2['M39'].value
sheet2['O41'] = sheet2['O39'].value
sheet2['Q41'] = sheet2['Q39'].value
sheet2['S41'] = sheet2['S39'].value
#----------------
sheet2['I42'] = sheet7['B21'].value
sheet2['K42'] = sheet7['C21'].value
sheet2['M42'] = sheet7['D21'].value
sheet2['O42'] = sheet7['E21'].value
sheet2['Q42'] = sheet7['F21'].value
sheet2['S42'] = sheet7['G21'].value
#----------------
sheet2['I43'] = sheet7['B22'].value
sheet2['K43'] = sheet7['C22'].value
sheet2['M43'] = sheet7['D22'].value
sheet2['O43'] = sheet7['E22'].value
sheet2['Q43'] = sheet7['F22'].value
sheet2['S43'] = sheet7['G22'].value
#----------------
sheet2['I44'] = sheet7['B23'].value
sheet2['K44'] = sheet7['C23'].value
sheet2['M44'] = sheet7['D23'].value
sheet2['O44'] = sheet7['E23'].value
sheet2['Q44'] = sheet7['F23'].value
sheet2['S44'] = sheet7['G23'].value
#----------------
sheet2['I45'] = 0
sheet2['K45'] = 0
sheet2['M45'] = 0
sheet2['O45'] = 0
sheet2['Q45'] = 0
sheet2['S45'] = 0
#----------------
sheet2['I46'] = sheet7['B23'].value
sheet2['K46'] = sheet7['C23'].value
sheet2['M46'] = sheet7['D23'].value
sheet2['O46'] = sheet7['E23'].value
sheet2['Q46'] = sheet7['F23'].value
sheet2['S46'] = sheet7['G23'].value
#----------------
sheet2['I47'] = sheet7['B37'].value
sheet2['K47'] = sheet7['C37'].value
sheet2['M47'] = sheet7['D37'].value
sheet2['O47'] = sheet7['E37'].value
sheet2['Q47'] = sheet7['F37'].value
sheet2['S47'] = sheet7['G37'].value
#----------------
sheet2['I48'] = sheet7['B38'].value
sheet2['K48'] = sheet7['C38'].value
sheet2['M48'] = sheet7['D38'].value
sheet2['O48'] = sheet7['E38'].value
sheet2['Q48'] = sheet7['F38'].value
sheet2['S48'] = sheet7['G38'].value
#----------------
sheet2['I49'] = sheet7['B39'].value
sheet2['K49'] = sheet7['C39'].value
sheet2['M49'] = sheet7['D39'].value
sheet2['O49'] = sheet7['E39'].value
sheet2['Q49'] = sheet7['F39'].value
sheet2['S49'] = sheet7['G39'].value
#----------------
sheet2['I50'] = 0
sheet2['K50'] = 0
sheet2['M50'] = 0
sheet2['O50'] = 0
sheet2['Q50'] = 0
sheet2['S50'] = 0
#----------------
sheet2['I51'] = 0
sheet2['K51'] = 0
sheet2['M51'] = 0
sheet2['O51'] = 0
sheet2['Q51'] = 0
sheet2['S51'] = 0
#----------------
sheet2['I52'] = sheet2['I49'].value
sheet2['K52'] = sheet2['K49'].value
sheet2['M52'] = sheet2['M49'].value
sheet2['O52'] = sheet2['O49'].value
sheet2['Q52'] = sheet2['Q49'].value
sheet2['S52'] = sheet2['S49'].value
#----------------
sheet2['I53'] = 0
sheet2['K53'] = 0
sheet2['M53'] = 0
sheet2['O53'] = 0
sheet2['Q53'] = 0
sheet2['S53'] = 0
#----------------
sheet2['I54'] = sheet2['I52'].value - sheet2['I53'].value
sheet2['K54'] = sheet2['K52'].value - sheet2['K53'].value
sheet2['M54'] = sheet2['M52'].value - sheet2['M53'].value
sheet2['O54'] = sheet2['O52'].value - sheet2['O53'].value
sheet2['Q54'] = sheet2['Q52'].value - sheet2['Q53'].value
sheet2['S54'] = sheet2['S52'].value - sheet2['S53'].value
#----------------
sheet2['I55'] = sheet7['B40'].value
sheet2['K55'] = sheet7['C40'].value
sheet2['M55'] = sheet7['D40'].value
sheet2['O55'] = sheet7['E40'].value
sheet2['Q55'] = sheet7['F40'].value
sheet2['S55'] = sheet7['G40'].value
#----------------
sheet2['I56'] = sheet7['B41'].value
sheet2['K56'] = sheet7['C41'].value
sheet2['M56'] = sheet7['D41'].value
sheet2['O56'] = sheet7['E41'].value
sheet2['Q56'] = sheet7['F41'].value
sheet2['S56'] = sheet7['G41'].value
#----------------
sheet2['I57'] = sheet7['B42'].value
sheet2['K57'] = sheet7['C42'].value
sheet2['M57'] = sheet7['D42'].value
sheet2['O57'] = sheet7['E42'].value
sheet2['Q57'] = sheet7['F42'].value
sheet2['S57'] = sheet7['G42'].value
#----------------
sheet2['I58'] = sheet7['B43'].value
sheet2['K58'] = sheet7['C43'].value
sheet2['M58'] = sheet7['D43'].value
sheet2['O58'] = sheet7['E43'].value
sheet2['Q58'] = sheet7['F43'].value
sheet2['S58'] = sheet7['G43'].value
#----------------
sheet2['I59'] = sheet7['B44'].value
sheet2['K59'] = sheet7['C44'].value
sheet2['M59'] = sheet7['D44'].value
sheet2['O59'] = sheet7['E44'].value
sheet2['Q59'] = sheet7['F44'].value
sheet2['S59'] = sheet7['G44'].value
#----------------
sheet2['I60'] = sheet7['B45'].value
sheet2['K60'] = sheet7['C45'].value
sheet2['M60'] = sheet7['D45'].value
sheet2['O60'] = sheet7['E45'].value
sheet2['Q60'] = sheet7['F45'].value
sheet2['S60'] = sheet7['G45'].value
#----------------
sheet2['I61'] = sheet7['B46'].value
sheet2['K61'] = sheet7['C46'].value
sheet2['M61'] = sheet7['D46'].value
sheet2['O61'] = sheet7['E46'].value
sheet2['Q61'] = sheet7['F46'].value
sheet2['S61'] = sheet7['G46'].value
#----------------
sheet2['I62'] = sheet7['B47'].value
sheet2['K62'] = sheet7['C47'].value
sheet2['M62'] = sheet7['D47'].value
sheet2['O62'] = sheet7['E47'].value
sheet2['Q62'] = sheet7['F47'].value
sheet2['S62'] = sheet7['G47'].value
#----------------
sheet2['I63'] = sheet7['B48'].value
sheet2['K63'] = sheet7['C48'].value
sheet2['M63'] = sheet7['D48'].value
sheet2['O63'] = sheet7['E48'].value
sheet2['Q63'] = sheet7['F48'].value
sheet2['S63'] = sheet7['G48'].value
#----------------
sheet2['I64'] = sheet7['B49'].value
sheet2['K64'] = sheet7['C49'].value
sheet2['M64'] = sheet7['D49'].value
sheet2['O64'] = sheet7['E49'].value
sheet2['Q64'] = sheet7['F49'].value
sheet2['S64'] = sheet7['G49'].value

wb2.save(OUTPUT_FILE)

print("*------全てのPLデータの転送が完了しました！------*")

#-----------------------------------------------------------------------#
pyautogui.alert('海外現法PLデータ転記を終了しました')#プログラム終了案内
#-----------------------------------------------------------------------#