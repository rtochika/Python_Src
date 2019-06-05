import csv
import openpyxl
import time
import datetime
from openpyxl.styles import Font
from openpyxl.styles import Alignment

#import pprint
#-------------------------------------------------#
# SMBC
#
#---------------------------------------------------------*
#本イニシャルファイルの説明
#
#---------------------------------------------------------*
#区分1,区分2,ファイル名,タブ名
#---------------------------------------------------------*
#1,1,C:\k-net\rpa\nyusyu.csv
#1,2,C:\k-net\rpa\仮受データ2018.xlsx,4月
#1,3,C:\k-net\rpa\\仮受明細1904.xlsx,4月

INITIAL_FILE = r'c:\k-net\rpa\initial_file.txt'

out_put_kariuke = ""  # 初期化
out_put_meisai = ""  # 初期化

A_COL = 0  # A列目の添字　 2:入金
C_COL = 2  #C列の添字：合計金額 A_COLが8のときだけ
D_COL = 3  #日付の添字 例："310303" → 0303
E_COL = 4  # ４列名の添字
F_COL = 5  # ５列名の添字
G_COL = 6  # ６列名の添字（入金額）
O_COL  =14  #入金額先
P_COL  =15  #銀行
Q_COL  =16  #銀行支店
R_COL  =17  #種別

#仮受ファイルフォーマット：余分な列は省略している
#------------------------------------#
with open(INITIAL_FILE) as f:
    rd = csv.reader(f)
    for rw in rd:
        if (rw[0] == "1"):  # 初期設定ファイルが記載されている
            if (rw[1] == "1"): #C:\k-net\rpa\nyusyu.csv
                INPUT_CSV=rw[2]
            elif (rw[1] == "2"): #C:\k-net\rpa\仮受デー仮タ2018.xlsx
                OUTPUT_FILE_EXCEL_KARIUKE=rw[2]
                OUTPUT_KARIUKE_TSUKI=rw[3]  #例：4月（タブ）指定
            elif (rw[1] == "3"): #C:\k-net\rpa\仮受明細1904.xlsx
                OUTPUT_FILE_EXCEL_MEISAI=rw[2]
                OUTPUT_MEISAI_TSUKI = rw[3]  # 例：4月（タブ）指定
#print(INPUT_CSV,OUTPUT_FILE_EXCEL_KARIUKE,OUTPUT_FILE_EXCEL_MEISAI)
#------------------------------------#

#仮受ファイル
wb_kar = openpyxl.load_workbook(OUTPUT_FILE_EXCEL_KARIUKE)  # 仮受ファイル
print(wb_kar.get_sheet_names())
sheet_kar=wb_kar.get_sheet_by_name(OUTPUT_KARIUKE_TSUKI)
#sheet_kar=wb_kar.get_sheet_by_name("4月")

#仮受明細ファイル
wb_mei = openpyxl.load_workbook(OUTPUT_FILE_EXCEL_MEISAI)  # 仮受明細ファイル
print(wb_mei.get_sheet_names())
sheet_mei=wb_mei.get_sheet_by_name(OUTPUT_MEISAI_TSUKI)
#sheet_mei=wb_mei.get_sheet_by_name("4月")

#init_proc-------------------------------------------------------------#
def init_proc(flg):
    #print("max: ",sheet_kar.max_row)
    if flg=='1':
        #return sheet_kar.max_row #最終行
        for cnt in range(2,1000):#２行目までヘッダー
            if sheet_kar['B'+str(cnt)].value:
                print(sheet_kar['C'+str(cnt)].value)
            else:
                break
        return cnt-1#1行マイナスする。

    elif flg=='2':
        for cnt in range(9,50):#８行目までヘッダー
            if sheet_mei['C'+str(cnt)].value:
                print(sheet_mei['C'+str(cnt)].value)
            else:
                break
        return cnt-1#1行マイナスする。
    elif flg=='3':#海外送金の最終行を知る
        for cnt in range(10,50):#９行目までヘッダー
            if sheet_mei['P'+str(cnt)].value:
                print(sheet_mei['P'+str(cnt)].value)
            else:
                break
        return cnt-1#1行マイナスする。
#end_proc-------------------------------------------------------------#
def end_proc():
    wb_kar.save(OUTPUT_FILE_EXCEL_KARIUKE) #仮受ファイル更新
    wb_mei.save(OUTPUT_FILE_EXCEL_MEISAI) # 仮受明細ファイル更新
#data_proc1:仮受処理-------------------------------------------------------------#
def data_proc1():
    kariuke_max_row =init_proc("1")#一次書込ファイルの最終行をゲット
    cntr=0
#仮受処理-------------------------------------------*
    with open(INPUT_CSV) as f:
        reader = csv.reader(f)
        for row in reader:
            if (row[A_COL] == "2"):  #入金データ　０列目区分=2
                if(row[E_COL]=="1"): #振込　４列目区分=1
                    cntr+=1
                    date=str(row[D_COL][-4:])
                    date=date[0:2]+"月"+date[2:4]+"日"
                    date=date.replace("0","") ##6桁から月日を抽出する（300303）→0303→3月3日
                    sheet_kar['A' + str(cntr + kariuke_max_row)].value = date
                    sheet_kar['B'+str(cntr+kariuke_max_row)].value=int(row[G_COL]) #入金額
                    sheet_kar['C'+str(cntr+kariuke_max_row)].value=row[O_COL] #入金額先
                    sheet_kar['D'+str(cntr+kariuke_max_row)].value=row[P_COL] #銀行
                    sheet_kar['E'+str(cntr+kariuke_max_row)].value=row[Q_COL] #銀行支店
                    sheet_kar['F'+str(cntr+kariuke_max_row)].value=row[R_COL] #種別
            elif (row[A_COL] == "8"): #０列目区分=8の時合計：
                cntr += 1
                sheet_kar['A' + str(cntr + kariuke_max_row)].alignment = Alignment(horizontal='right')
                sheet_kar['A' + str(cntr + kariuke_max_row)].value = "＊合計*" # 今日の合計
                #sheet_kar['A' + str(cntr + kariuke_max_row)].font = Font(bold = True)
                sheet_kar['B' + str(cntr + kariuke_max_row)].value = int(row[C_COL])  # 合計金額
                date = datetime.datetime.now()
                date = str(date.hour)+":"+str(date.minute)
                sheet_kar['G' + str(cntr-1 + kariuke_max_row)].value = date #データ取得時間
                sheet_kar['H' + str(cntr-1 + kariuke_max_row)].value = str(cntr-1)+" 件" #件数

#data_proc2:仮受明細処理-------------------------------------------------------------#
def data_proc2():
#---スペース行の確認-------"
    karimei_max_row = init_proc("2")  # ２次書込ファイル（仮受明細）の最終行をゲット
    karimei_max_kaigai=init_proc("3")  # ２次書込ファイル（仮受明細-海外分）の最終行をゲット
    cntr=0 #仮受明細のカウンター
    cntr_k=0 #海外のカウンター
    gokei =0 #合計金額
#仮受明細処理-------------------------------------------*
    with open(INPUT_CSV) as f:
        reader = csv.reader(f)
        for row in reader:
            if (row[A_COL] == "2"):  #入金データ　０列目区分=2
                if(row[E_COL]=="1"): #振込　４列目区分=1
                    if (row[F_COL] == "11"):  # 振込　５列目区分=11
                        cntr+=1
                        date = str(row[D_COL][-4:])
                        date = date[0:2] + "月" + date[2:4] + "日"
                        date=date.replace("0","") ##6桁から月日を抽出する（300303）→0303→3月3日
                        sheet_mei['E' + str(cntr + karimei_max_row)].value = date
                        sheet_mei['C'+str(cntr+karimei_max_row)].value=int(row[G_COL]) #入金額
                        gokei=gokei+int(row[G_COL])#合計金額サムアップ
                        sheet_mei['D'+str(cntr+karimei_max_row)].value=row[O_COL] #入金額先
                    elif (row[F_COL] == "14"):  # 振込　５列目区分=14は海外振込み
                        cntr_k+=1
                        sheet_mei['P' + str(cntr_k + karimei_max_kaigai)].value = int(row[G_COL])  # 入金額
                        sheet_mei['U' + str(cntr_k + karimei_max_kaigai)].value = int(row[G_COL])  # 入金額
                        #print("海外：", row[G_COL])

            elif (row[A_COL] == "8"): #０列目区分=8の時合計：
                #sheet_mei['A' + str(cntr + karimei_max_row)].font = Font(bold = True)
                #sheet_mei['A' + str(cntr + karimei_max_row)].value = int(row[C_COL])  # 合計金額
                sheet_mei['A' + str(cntr + karimei_max_row)].value = gokei

#処理スタート-------------------------#
data_proc1()#仮受処理
data_proc2()#仮受明細処理
end_proc()#後処理

