import time
from selenium import webdriver
import openpyxl

deb_flg =0 # 1:ON 0:OFF
#【機能】---------------------------------------------------*
#海外現地法人の収支実績を転記する
# インプット・アウトプット対比表 → http://bit.ly/2XDFG9J
#-----------------------------------------------------------*

#ファイル名処理-----------------------------------------------------------*
f = open('kaig_ten_input_file.txt','r', encoding='utf-8')
data1 = f.read()  # ファイル終端まで全て読んだデータを返す
f.close()
print(type(data1)) # 文字列データ
lines1 = data1.split('\n') # 改行で区切る(改行文字そのものは戻り値のデータには含まれない)
print(type(lines1))
cntr=0
for line in lines1:#　#がある行は読み飛ばす
    idx = line.find("#")
    if idx != 0:#行の１文字目に#がないものがファイル名
        cntr = cntr + 1
        #print(line)
        if cntr==1:
            OUTPUT_FILE = line #例：海外18012.xlsx
        elif cntr==2:
            SNG_IN_F = line #KMTS2018.12.xlsx
        elif cntr==3:
            HNK_IN_F = line #KMTH2018.12.xlsx
        elif cntr==4:
            SHA_IN_F = line #KMT上海2018.12.xlsx
        elif cntr == 5:
            PHI_IN_F = line #KEFCI2018.12 .xlsx
        elif cntr == 6:
            TWN_IN_F = line #KMTT2018.12.xlsx
        elif cntr == 7:
            VTN_IN_F = line #KMTV2018.12.xlsx

if deb_flg == 0:
    print(OUTPUT_FILE)
    print(SNG_IN_F)
    print(HNK_IN_F)
    print(SHA_IN_F)
    print(PHI_IN_F)
    print(TWN_IN_F)
    print(VTN_IN_F)


#ファイル名
#------------------------------------------------
#OUTPUT_FILE = "海外18012.xlsx"
#SNG_IN_F = "【C】KMTS2018.12.xlsx"
#HNK_IN_F = "【C】KMTH2018.12.xlsx"
#SHA_IN_F = "【C】KMT上海2018.12.xlsx"
#PHI_IN_F = "【C】KEFCI2018.12 .xlsx"
#TWN_IN_F = "【C】KMTT2018.12.xlsx"
#VTN_IN_F = "【C】KMTV2018.12.xlsx"
#------------------------------------------------

from selenium.webdriver.remote.webelement import WebElement

#閲覧文書終始実績Excelデータ操作
#wb2 = openpyxl.load_workbook("海外18012.xlsx")
wb2 = openpyxl.load_workbook(OUTPUT_FILE.strip())

#上海----------------------------------------------------*
sheet2=wb2.get_sheet_by_name("SH")#上海（アウトプット）
#wb_sh = openpyxl.load_workbook("【C】KMT上海2018.12.xlsx", data_only=True)
wb_sh = openpyxl.load_workbook(SHA_IN_F.strip(), data_only=True)

print(wb_sh.get_sheet_names())
print("*------上海開始！------*")
sheet4=wb_sh.get_sheet_by_name("3.PL")#上海PL（インプット）
for cnt in range(9,48+1):
    if(cnt<=28):
        sheet2['I'+str(cnt)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt)]=sheet4['P'+str(cnt)].value
    #elif(cnt>=34 and cnt <=41):
    elif (cnt >= 34 and cnt <= 42):
        print("I:", (cnt+10))
        sheet2['I'+str(cnt+10)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt+10)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt+10)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt+10)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt+10)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt+10)]=sheet4['P'+str(cnt)].value
    elif(cnt==43):
        print("I:", (cnt+10))
        sheet2['I'+str(cnt+10)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt+10)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt+10)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt+10)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt+10)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt+10)]=sheet4['P'+str(cnt)].value
    elif (cnt >=46):
        print("I*", (cnt+10))
        sheet2['I'+str(cnt+10)]=sheet4['K'+str(cnt)].value
        sheet2['K'+str(cnt+10)]=sheet4['L'+str(cnt)].value
        sheet2['M'+str(cnt+10)]=sheet4['M'+str(cnt)].value
        sheet2['O'+str(cnt+10)]=sheet4['N'+str(cnt)].value
        sheet2['Q'+str(cnt+10)]=sheet4['O'+str(cnt)].value
        sheet2['S'+str(cnt+10)]=sheet4['P'+str(cnt)].value
print("I52",sheet2['I52'].value)

wb2.save('海外18012.xlsx')

